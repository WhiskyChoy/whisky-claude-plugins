#!/usr/bin/env python3
"""
Generate formatted XLSX review files from correction data.

Usage:
    python generate_review_xlsx.py glossary --input data.json --output glossary.xlsx
    python generate_review_xlsx.py glossary --input merged.json --key glossary --output glossary.xlsx
    python generate_review_xlsx.py glossary --input merged.json --key glossary \
        --speaker-map speaker-map.json --transcript transcript.json --output glossary.xlsx
    python generate_review_xlsx.py corrections --input data.json --output corrections.xlsx
    python generate_review_xlsx.py commitments --input data.json --output commitments.xlsx

Input JSON: either a flat array [{...}, ...] or a nested object {"key": [{...}, ...]}.
Use --key to extract from nested objects. If omitted and input is a dict,
auto-detects by matching the subcommand name (e.g. "glossary" key for glossary command).

Glossary JSON row format:
    {"类别": "术语", "原文/变体": "...", "修正": "...", "出现次数": 10, "示例上下文": "..."}

Corrections JSON row format:
    {"时间戳": "00:01:26", "说话人": "SPEAKER_01", "原文": "...", "修正": "...", "置信度": 0.95, "类型": "..."}
"""

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Style definitions ──────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# Category row fills (glossary) — order also defines sort priority
CATEGORY_ORDER = ["术语", "人名", "说话人", "口语简化"]
CATEGORY_FILLS = {
    "术语": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "人名": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "说话人": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "口语简化": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

DATA_ROW_FILL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")


# ── Helpers ────────────────────────────────────────────────────────────────

def _estimate_col_width(values, max_width=60):
    """Estimate column width from content (CJK chars count as 2)."""
    max_len = 0
    for val in values:
        char_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
        max_len = max(max_len, min(char_len, max_width))
    return max(max_len + 4, 10)


def _apply_common_formatting(ws, rows, category_col=None):
    """Apply header, borders, column widths, freeze panes, auto-filter."""
    if not rows:
        return

    num_cols = len(rows[0])

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN

            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
            elif category_col is not None and category_col < len(row):
                category = row[category_col]
                fill = CATEGORY_FILLS.get(category)
                if fill:
                    cell.fill = fill
            else:
                has_content = any(str(v).strip() for v in row if v)
                if has_content:
                    cell.fill = DATA_ROW_FILL

    for col_idx in range(1, num_cols + 1):
        col_values = [row[col_idx - 1] for row in rows if col_idx <= len(row)]
        ws.column_dimensions[get_column_letter(col_idx)].width = _estimate_col_width(col_values)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _sort_glossary_rows(rows):
    """Sort glossary data rows by category order, then by count descending.
    rows[0] is the header row (preserved). Remaining rows are sorted."""
    if len(rows) <= 1:
        return rows

    header = rows[0]
    data = rows[1:]

    # Find column indices
    cat_col = next((i for i, h in enumerate(header) if "类别" in str(h)), 0)
    count_col = next((i for i, h in enumerate(header) if "次数" in str(h)), 3)

    cat_priority = {c: i for i, c in enumerate(CATEGORY_ORDER)}

    def sort_key(row):
        cat = str(row[cat_col]) if cat_col < len(row) else ""
        count = row[count_col] if count_col < len(row) else 0
        try:
            count = int(count)
        except (ValueError, TypeError):
            count = 0
        return (cat_priority.get(cat, 99), -count)

    data.sort(key=sort_key)
    return [header] + data


# ── XLSX creators ──────────────────────────────────────────────────────────

def create_glossary_xlsx(rows, output_path):
    """Create a formatted glossary XLSX. rows[0] is the header.
    Auto-sorts by category group then count descending."""
    rows = _sort_glossary_rows(rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全局映射表"
    _apply_common_formatting(ws, rows, category_col=0)
    wb.save(output_path)
    print(f"Glossary: {output_path} ({len(rows) - 1} rows)")


def create_corrections_xlsx(rows, output_path):
    """Create a formatted corrections XLSX. rows[0] is the header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "逐条修正"
    _apply_common_formatting(ws, rows, category_col=None)
    wb.save(output_path)
    print(f"Corrections: {output_path} ({len(rows) - 1} rows)")


def create_commitments_xlsx(rows, output_path):
    """Create a formatted commitments tracking XLSX. rows[0] is the header.
    Color-codes by commitment strength (承诺强度) column."""
    STRENGTH_FILLS = {
        "明确承诺": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "被分配": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "条件承诺": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "被分配（未确认）": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "建议": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }
    WARNING_BORDER = Border(
        left=Side(style="medium", color="FF0000"),
        right=Side(style="medium", color="FF0000"),
        top=Side(style="medium", color="FF0000"),
        bottom=Side(style="medium", color="FF0000"),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "承诺追踪"

    if not rows:
        wb.save(output_path)
        return

    headers = [str(c).strip() if c else "" for c in rows[0]]
    strength_col = next((i for i, h in enumerate(headers) if "强度" in h), None)
    risk_col = next((i for i, h in enumerate(headers) if "风险" in h or "⚠" in h), None)

    num_cols = len(rows[0])
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN

            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
            else:
                if strength_col is not None and strength_col < len(row):
                    fill = STRENGTH_FILLS.get(str(row[strength_col]).strip())
                    if fill:
                        cell.fill = fill
                if risk_col is not None and risk_col < len(row):
                    risk_val = str(row[risk_col]).strip().lower()
                    if risk_val and risk_val not in ("false", "0", "", "none", "无"):
                        cell.border = WARNING_BORDER

    for col_idx in range(1, num_cols + 1):
        col_values = [row[col_idx - 1] for row in rows if col_idx <= len(row)]
        ws.column_dimensions[get_column_letter(col_idx)].width = _estimate_col_width(col_values)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    print(f"Commitments: {output_path} ({len(rows) - 1} rows)")


# ── JSON/TSV → XLSX conversion ────────────────────────────────────────────

def _load_json_data(json_path, key=None, sheet_type=None):
    """Load JSON, supporting both flat arrays and nested objects.
    - If key is given, extract data[key]
    - If data is a dict and no key given, try data[sheet_type]
    - If data is a list, use directly
    """
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    if isinstance(raw, list):
        return raw

    if isinstance(raw, dict):
        if key and key in raw:
            return raw[key]
        if sheet_type and sheet_type in raw:
            return raw[sheet_type]
        # Try common key names
        for k in ["glossary", "corrections", "commitments", "data", "items"]:
            if k in raw:
                return raw[k]
        raise ValueError(f"Cannot extract array from JSON dict. Keys: {list(raw.keys())}. Use --key to specify.")

    raise ValueError(f"Unexpected JSON type: {type(raw)}")


def _inject_speaker_rows(data, speaker_map_path, transcript_path=None):
    """Append speaker mapping rows to glossary data from speaker-map.json."""
    with open(speaker_map_path, "r", encoding="utf-8") as f:
        spk_map = json.load(f)

    # Get segment counts from transcript if available
    seg_counts = {}
    if transcript_path and Path(transcript_path).exists():
        with open(transcript_path, "r", encoding="utf-8") as f:
            transcript = json.load(f)
        for seg in transcript.get("segments", []):
            spk = seg.get("speaker", "")
            seg_counts[spk] = seg_counts.get(spk, 0) + 1

    # Detect field names: could be Chinese or English
    if data:
        sample = data[0]
        cat_key = "category" if "category" in sample else "类别"
        orig_key = "originals" if "originals" in sample else "原文/变体"
        corr_key = "correction" if "correction" in sample else "修正"
        count_key = "count" if "count" in sample else "出现次数"
        example_key = "example" if "example" in sample else "示例上下文"
    else:
        cat_key, orig_key, corr_key, count_key, example_key = \
            "category", "originals", "correction", "count", "example"

    for spk_id, name in sorted(spk_map.items(), key=lambda x: -seg_counts.get(x[0], 0)):
        data.append({
            cat_key: "说话人",
            orig_key: spk_id,
            corr_key: name,
            count_key: seg_counts.get(spk_id, 0),
            example_key: "",
        })

    return data


def _inject_filler_rows(data, fillers_path):
    """Append filler/simplification rows to glossary data from a JSON file."""
    with open(fillers_path, "r", encoding="utf-8") as f:
        fillers = json.load(f)

    if data:
        keys = list(data[0].keys())
    else:
        keys = ["类别", "原文/变体", "修正", "出现次数", "示例上下文"]

    for filler in fillers:
        row = {keys[i]: filler.get(k, "") for i, k in enumerate(["类别", "原文/变体", "修正", "出现次数", "示例上下文"]) if i < len(keys)}
        row[keys[0]] = "口语简化"
        data.append(row)

    return data


def json_to_xlsx(json_path, xlsx_path, sheet_type="glossary", key=None,
                 speaker_map=None, transcript=None, fillers=None):
    """Convert JSON data to formatted XLSX."""
    data = _load_json_data(json_path, key=key, sheet_type=sheet_type)

    if not data:
        print(f"Empty data from: {json_path}")
        return

    # Inject speaker and filler rows for glossary
    if sheet_type == "glossary":
        if speaker_map:
            data = _inject_speaker_rows(data, speaker_map, transcript)
        if fillers:
            data = _inject_filler_rows(data, fillers)

    # Normalize field names: map English keys to Chinese display headers
    _FIELD_MAP = {
        # glossary
        "category": "类别", "originals": "原文/变体", "correction": "修正",
        "count": "出现次数", "example": "示例上下文",
        # corrections
        "timestamp": "时间戳", "speaker": "说话人", "original": "原文",
        "corrected": "修正", "confidence": "置信度", "reason": "类型",
    }
    # Expected column order per sheet type
    _COL_ORDER = {
        "glossary": ["类别", "原文/变体", "修正", "出现次数", "示例上下文"],
        "corrections": ["时间戳", "说话人", "原文", "修正", "置信度", "类型"],
        "commitments": None,  # use as-is
    }

    def _cell_value(val):
        if isinstance(val, list):
            return " / ".join(str(v) for v in val)
        return val

    # Map keys
    raw_keys = list(data[0].keys())
    mapped_keys = [_FIELD_MAP.get(k, k) for k in raw_keys]

    # Determine column order
    ordered = _COL_ORDER.get(sheet_type)
    if ordered:
        # Build index: display_name → raw_key
        name_to_raw = {_FIELD_MAP.get(k, k): k for k in raw_keys}
        # Use ordered columns that exist, append any extras
        final_cols = [(h, name_to_raw[h]) for h in ordered if h in name_to_raw]
        seen = {h for h, _ in final_cols}
        for display, raw in zip(mapped_keys, raw_keys):
            if display not in seen:
                final_cols.append((display, raw))
        headers = [h for h, _ in final_cols]
        raw_order = [r for _, r in final_cols]
    else:
        headers = mapped_keys
        raw_order = raw_keys

    rows = [headers] + [[_cell_value(item.get(r, "")) for r in raw_order] for item in data]

    creators = {
        "glossary": create_glossary_xlsx,
        "corrections": create_corrections_xlsx,
        "commitments": create_commitments_xlsx,
    }
    creators.get(sheet_type, create_corrections_xlsx)(rows, xlsx_path)


def convert_tsv_to_xlsx(tsv_path, xlsx_path, sheet_type="glossary"):
    """Convert a TSV file (with optional BOM) to formatted XLSX."""
    rows = []
    with open(tsv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter="\t")
        for row in reader:
            rows.append(row)

    if not rows:
        print(f"Empty TSV: {tsv_path}")
        return

    creators = {"glossary": create_glossary_xlsx, "corrections": create_corrections_xlsx}
    creators.get(sheet_type, create_corrections_xlsx)(rows, xlsx_path)


# ── CLI ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate formatted XLSX review files")
    subparsers = parser.add_subparsers(dest="command", required=True)

    # glossary subcommand
    p_glossary = subparsers.add_parser("glossary", help="Generate glossary XLSX from JSON")
    p_glossary.add_argument("--input", required=True, help="Input JSON file (flat array or nested object)")
    p_glossary.add_argument("--output", required=True, help="Output XLSX file")
    p_glossary.add_argument("--key", default=None, help="Key to extract from nested JSON (default: auto-detect)")
    p_glossary.add_argument("--speaker-map", default=None, help="Speaker-map JSON to inject as 说话人 rows")
    p_glossary.add_argument("--transcript", default=None, help="Transcript JSON (for speaker segment counts)")
    p_glossary.add_argument("--fillers", default=None, help="Filler words JSON to inject as 口语简化 rows")

    # corrections subcommand
    p_corrections = subparsers.add_parser("corrections", help="Generate corrections XLSX from JSON")
    p_corrections.add_argument("--input", required=True, help="Input JSON file")
    p_corrections.add_argument("--output", required=True, help="Output XLSX file")
    p_corrections.add_argument("--key", default=None, help="Key to extract from nested JSON")

    # commitments subcommand
    p_commitments = subparsers.add_parser("commitments", help="Generate commitments tracking XLSX from JSON")
    p_commitments.add_argument("--input", required=True, help="Input JSON file")
    p_commitments.add_argument("--output", required=True, help="Output XLSX file")
    p_commitments.add_argument("--key", default=None, help="Key to extract from nested JSON")

    # convert subcommand (TSV → XLSX)
    p_convert = subparsers.add_parser("convert", help="Convert TSV to formatted XLSX")
    p_convert.add_argument("--input", required=True, help="Input TSV file")
    p_convert.add_argument("--output", required=True, help="Output XLSX file")
    p_convert.add_argument("--sheet-type", choices=["glossary", "corrections", "commitments"], default="glossary")

    args = parser.parse_args()

    if args.command == "convert":
        convert_tsv_to_xlsx(args.input, args.output, args.sheet_type)
    elif args.command == "glossary":
        json_to_xlsx(args.input, args.output, "glossary",
                     key=args.key, speaker_map=args.speaker_map,
                     transcript=args.transcript, fillers=args.fillers)
    elif args.command == "corrections":
        json_to_xlsx(args.input, args.output, "corrections", key=getattr(args, "key", None))
    elif args.command == "commitments":
        json_to_xlsx(args.input, args.output, "commitments", key=getattr(args, "key", None))


if __name__ == "__main__":
    main()
